import os
import pandas as pd
import sqlite3
from datetime import datetime
import imaplib
import email
from email.header import decode_header
import re
from dotenv import load_dotenv
import numpy as np
from openpyxl import load_workbook
import streamlit as st
import locale

# Загрузка переменных окружения
load_dotenv()
DB_PATH = os.getenv("DB_PATH", "db.db")


def get_email_credentials():
    email_user = os.getenv("EMAIL_USER")
    email_password = os.getenv("EMAIL_PASSWORD")
    email_server = os.getenv("EMAIL_SERVER", "imap.gmail.com")
    if not all([email_user, email_password, email_server]):
        raise RuntimeError("Не найдены переменные окружения EMAIL_USER, EMAIL_PASSWORD или EMAIL_SERVER")
    return email_user, email_password, email_server


def connect_to_email():
    user, password, server = get_email_credentials()
    mail = imaplib.IMAP4_SSL(server)
    mail.login(user, password)
    return mail


def get_latest_stock_excel():
    mail = connect_to_email()
    mail.select('inbox')
    _, messages = mail.search(None, 'SINCE', (datetime.now() - pd.Timedelta(days=14)).strftime("%d-%b-%Y"))
    if not messages[0]:
        return None, None

    for email_id in messages[0].split()[::-1][:10]:
        _, msg_data = mail.fetch(email_id, '(RFC822)')
        msg = email.message_from_bytes(msg_data[0][1])
        subject, encoding = decode_header(msg['Subject'])[0]
        if isinstance(subject, bytes):
            subject = subject.decode(encoding or 'utf-8')
        if "ведомость" in subject.lower() or "склад на от" in subject.lower():
            for part in msg.walk():
                if part.get_content_maintype() == 'multipart':
                    continue
                if not part.get('Content-Disposition'):
                    continue
                filename, enc = decode_header(part.get_filename())[0]
                if isinstance(filename, bytes):
                    filename = filename.decode(enc or 'utf-8')
                if filename.lower().endswith(('.xlsx', '.xls')):
                    os.makedirs('temp', exist_ok=True)
                    path = os.path.join('temp', filename)
                    with open(path, 'wb') as f:
                        f.write(part.get_payload(decode=True))

                    body = ""
                    if msg.is_multipart():
                        for p in msg.walk():
                            if p.get_content_type() == "text/plain":
                                body = p.get_payload(decode=True).decode(p.get_content_charset() or 'utf-8', 'ignore')
                                break
                    else:
                        body = msg.get_payload(decode=True).decode(msg.get_content_charset() or 'utf-8', 'ignore')

                    m = re.search(r"(\d{1,2} \w+ \d{4} г\.)", body)
                    report_date = m.group(1) if m else datetime.now().strftime("%Y-%m-%d")
                    return path, report_date
    return None, None


def process_stock_excel(filepath, report_date=None):
    warehouse_map = {
        'СТН Склад ТМЦ': 'Склад (материалы)',
        'СТН Кладовая уч.кабельных изделий': 'Склад (материалы)',
        'Склад МОСКВА (ООО СТН)': 'Склад Готовой продукции',
        'Склад ВЛАДИМИР (Автоприбор)': 'Склад Готовой продукции',
        'СТН Склад гот.изделий': 'Склад Готовой продукции',
        'Склад НТЗ': 'Склад НТЗ'
    }

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    data = [row for row in ws.iter_rows(min_row=2, values_only=True)]
    df = pd.DataFrame(data, columns=[
        'Артикул', 'Номенклатура', 'Номенклатура.Вид номенклатуры',
        'Склад', 'Количество конечный остаток', 'Оценка (конечный остаток)'
    ])

    df.rename(columns={
        'Артикул': 'article',
        'Номенклатура': 'nomenclature',
        'Номенклатура.Вид номенклатуры': 'nomenclature_type',
        'Склад': 'warehouse',
        'Количество конечный остаток': 'quantity',
        'Оценка (конечный остаток)': 'value'
    }, inplace=True)

    df['warehouse'] = df['warehouse'].map(lambda x: warehouse_map.get(str(x).strip(), x))
    df['quantity'] = pd.to_numeric(df['quantity'], errors='coerce').fillna(0)
    df['value'] = pd.to_numeric(df['value'], errors='coerce').fillna(0)
    df['date_updated'] = datetime.now()
    df['report_date'] = report_date or datetime.now().strftime('%Y-%m-%d')
    return df


def update_stock_db(df):
    if df is None or df.empty:
        return False
    try:
        conn = sqlite3.connect(DB_PATH)
        df.to_sql('stock_balance', conn, if_exists='append', index=False)
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Ошибка при обновлении базы данных: {e}")
        return False


def update_stock_data():
    excel_file, report_date = get_latest_stock_excel()
    if not excel_file:
        print("Не найдено писем с Excel-файлами")
        return False

    df = process_stock_excel(excel_file, report_date)
    if df is None:
        return False

    success = update_stock_db(df)
    try:
        os.remove(excel_file)
    except:
        pass
    return success


def show_stale_stock():
    st.header("🧱 Залежавшийся складской остаток")
    try:
        locale.setlocale(locale.LC_TIME, 'ru_RU.UTF-8')
    except locale.Error:
        locale.setlocale(locale.LC_TIME, '')

    update_stock_data()

    conn = sqlite3.connect(DB_PATH)
    dates_df = pd.read_sql_query("SELECT DISTINCT report_date FROM stock_balance ORDER BY report_date DESC", conn)
    conn.close()

    dates_df['report_date'] = pd.to_datetime(dates_df['report_date'])
    formatted_labels = dates_df['report_date'].dt.strftime('%-d %B %Y').str.capitalize()
    label_to_date = dict(zip(formatted_labels, dates_df['report_date']))

    selected_label = st.selectbox("Выберите дату отчета:", list(label_to_date.keys()))
    selected_date = label_to_date[selected_label]

    st.markdown(f"📅 **Отображаются данные за дату: {selected_label}**")

    conn = sqlite3.connect(DB_PATH)
    query = """
    SELECT article, nomenclature, warehouse, quantity, report_date
    FROM stock_balance
    WHERE quantity > 0 AND report_date = ?
    """
    df = pd.read_sql_query(query, conn, params=(selected_date.strftime('%Y-%m-%d'),))
    conn.close()

    df['report_date'] = pd.to_datetime(df['report_date'])
    current_date = df['report_date'].max()

    df_first = (
        df.groupby(['article', 'warehouse'])
        .agg({
            'report_date': 'min',
            'nomenclature': 'first',
            'quantity': 'sum'
        })
        .reset_index()
    )
    df_first['storage_months'] = ((current_date - df_first['report_date']) / pd.Timedelta(days=30)).astype(int)

    def highlight_row(row):
        return [
            'background-color: #f8d7da; color: #721c24;' if row['storage_months'] >= 9 else ''
        ] * len(row)

    df_show = df_first[['article', 'nomenclature', 'warehouse', 'quantity', 'report_date', 'storage_months']]
    styled = df_show.style.apply(highlight_row, axis=1)
    st.dataframe(styled, use_container_width=True)
