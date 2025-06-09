import os
import pandas as pd
import sqlite3
from datetime import datetime
import imaplib
import email
from email.header import decode_header
import tempfile
from openpyxl import load_workbook
import logging
from dotenv import load_dotenv

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# ------------------ Почтовые утилиты ------------------
def get_email_credentials():
    user = os.getenv("EMAIL_USER")
    pwd  = os.getenv("EMAIL_PASSWORD")
    srv  = os.getenv("EMAIL_SERVER", "imap.gmail.com")
    missing = [v for v in ("EMAIL_USER","EMAIL_PASSWORD","EMAIL_SERVER") if not os.getenv(v)]
    if missing:
        raise RuntimeError(f"Не найдены переменные окружения: {', '.join(missing)}.")
    return user, pwd, srv

def connect_to_email():
    user, pwd, srv = get_email_credentials()
    mail = imaplib.IMAP4_SSL(srv)
    mail.login(user, pwd)
    return mail

def extract_date_from_subject(subject):
    try:
        for part in subject.split():
            try:
                return datetime.strptime(part, "%d.%m.%Y").strftime("%Y-%m-%d")
            except ValueError:
                continue
    except Exception:
        pass
    return datetime.now().strftime("%Y-%m-%d")

def decode_mime_words(s):
    decoded_fragments = decode_header(s)
    return ''.join([
        fragment.decode(encoding or 'utf-8') if isinstance(fragment, bytes) else fragment
        for fragment, encoding in decoded_fragments
    ])

def find_latest_purchase_file():
    try:
        logging.info("Подключение к почте...")
        mail = connect_to_email()
        logging.info("✅ Подключение к почте успешно")

        mail.select("inbox")
        logging.info("🔍 Поиск писем с темой, содержащей 'закупк'...")
        result, data = mail.search(None, "ALL")
        if result != "OK":
            logging.warning("❌ Не удалось получить письма")
            return None, None

        email_ids = data[0].split()[::-1][:10]
        logging.info(f"📨 Проверяем последние {len(email_ids)} писем")

        for eid in email_ids:
            result, msg_data = mail.fetch(eid, "(RFC822)")
            if result != "OK":
                continue

            msg = email.message_from_bytes(msg_data[0][1])
            subject_raw = msg.get("Subject", "")
            subject = decode_mime_words(subject_raw)

            if "закупк" not in subject.lower():
                continue

            logging.info(f"📦 Найдено письмо с темой: {subject}")

            for part in msg.walk():
                if part.get_content_maintype() == "multipart":
                    continue
                if part.get("Content-Disposition") is None:
                    continue

                filename_raw = part.get_filename()
                if not filename_raw:
                    continue
                filename = decode_mime_words(filename_raw)

                if filename.lower().endswith((".xls", ".xlsx")):
                    logging.info(f"⬇️ Скачивание вложения: {filename}")
                    payload = part.get_payload(decode=True)
                    if not payload:
                        continue
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                        tmp.write(payload)
                        return tmp.name, subject

        logging.warning("📭 Не найден подходящий файл вложения в письмах")
        return None, None

    except Exception as e:
        logging.error(f"❌ Ошибка в find_latest_purchase_file: {e}")
        return None, None

def update_purchases_data():
    DB_PATH = r"C:\\Users\\user\\Desktop\\Проекты\\projects\\sales_airflow_project\\db.db"

    try:
        logging.info("Поиск последнего файла закупок...")
        path, subj = find_latest_purchase_file()
        if not path:
            logging.warning("📭 Файл закупок не найден")
            return False

        report_date = extract_date_from_subject(subj)
        wb = load_workbook(path, data_only=True)
        ws = wb.active

        header = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
        col_map = {
            'поставщик': 1,
            'номенклатура': 0,
            'кол-во': 2,
            'цена': 3,
            'сумма': 4,
            'сумма с ндс': 5
        }

        data = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            supplier = row[col_map['поставщик']]
            product = row[col_map['номенклатура']]
            quantity = row[col_map['кол-во']]
            price_per_unit = row[col_map['цена']]
            total = row[col_map['сумма']]
            total_with_vat = row[col_map['сумма с ндс']]

            if not supplier or not product or not isinstance(quantity, (int, float)):
                continue

            data.append({
                "supplier": str(supplier).strip(),
                "product": str(product).strip(),
                "quantity": quantity,
                "price_per_unit": price_per_unit or 0,
                "total": total or 0,
                "total_with_vat": total_with_vat or 0,
                "report_date": report_date
            })

        if not data:
            logging.warning("❌ Нет данных для обновления закупок")
            return False

        df = pd.DataFrame(data)
        logging.info("🔧 Создание таблицы purchases при необходимости...")
        conn = sqlite3.connect(DB_PATH)
        conn.execute('''CREATE TABLE IF NOT EXISTS purchases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            supplier TEXT,
            product TEXT,
            quantity REAL,
            price_per_unit REAL,
            total REAL,
            total_with_vat REAL,
            report_date TEXT
        )''')
        logging.info("📥 Добавление данных...")
        df.to_sql('purchases', conn, if_exists='append', index=False)
        conn.commit()
        conn.close()
        os.remove(path)
        logging.info("✅ Закупки обновлены")
        return True

    except Exception as e:
        logging.error(f"❌ Ошибка при обновлении закупок: {e}")
        return False

if __name__ == "__main__":
    update_purchases_data()
